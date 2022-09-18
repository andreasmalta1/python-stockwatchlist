from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import yfinance as yf


def main():
    wb = load_workbook(filename="Stocks Watchlist.xlsx")
    ws = wb["Sheet1"]
    l_row = len(ws["A"])
    ticker = []
    skip_rows = []
    for row in range(3, l_row + 1):
        if ws["A" + str(row)].value == "CCTrader" or ws["A" + str(row)].value == "Watchlist":
            ticker.append(None)
            skip_rows.append(row)
        else:
            ticker.append(yf.Ticker(ws["A" + str(row)].value))

    list_info = ["regularMarketPrice", "fiftyTwoWeekLow", "fiftyTwoWeekHigh", "marketCap", "Short Long Term Debt",
                 "revenue", "netIncomeToCommon", "trailingEps", "assets", "liabilities", "dividendRate", "sector"]
    information = {}

    for stock in ticker:
        if stock:
            information[stock] = {}
            for item in list_info:
                try:
                    information[stock][item] = stock.info[item]
                except KeyError:
                    information[stock][item] = None

    skip_columns = [6, 14, 15, 17]
    col = 3
    row = 3
    for stock_id, stock_info in information.items():
        for key in stock_info:
            while col in skip_columns:
                col += 1
            char = get_column_letter(col)
            ws[char + str(row)] = stock_info[key]
            col += 1
        row += 1
        while row in skip_rows:
            row += 1
        col = 3

    wb.save("Stocks Watchlist.xlsx")


main()
